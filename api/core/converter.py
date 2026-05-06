# -*- coding: utf-8 -*-
"""
Conversion d'un fichier Excel SSSOM vers le format SSSOM/TSV.
Reproduit la logique de convert_to_sssom.py sous forme de module réutilisable.
"""

import io
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


class ConversionError(Exception):
    """Erreur levée si le classeur Excel ne respecte pas la structure attendue."""


def convert_xlsx_to_sssom(
    xlsx_bytes: bytes,
    output_path: str | Path,
) -> dict:
    """
    Lit un classeur Excel SSSOM et écrit le fichier .sssom.tsv correspondant.

    Le classeur doit contenir deux feuilles :
      - SSSOM_Metadata : deux colonnes (clé, valeur) à partir de la ligne 2.
      - Mappings       : en-tête en ligne 1, données à partir de la ligne 2.
                         Les lignes dont subject_id ne commence pas par
                         "mainframe:" sont ignorées (en-têtes de section).

    Retourne un dict de résumé :
      {
        "output_path": str,
        "metadata_entries": int,
        "curie_prefixes": int,
        "mapping_rows": int,
        "preview": [ {subject_id, predicate_id, object_id}, ... ]  (5 premiers)
      }
    """
    # ── Chargement ────────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        raise ConversionError(f"Impossible d'ouvrir le fichier Excel : {e}") from e

    for sheet_name in ("SSSOM_Metadata", "Mappings"):
        if sheet_name not in wb.sheetnames:
            raise ConversionError(
                f"Feuille '{sheet_name}' introuvable. "
                f"Feuilles présentes : {wb.sheetnames}"
            )

    # ── Lecture des métadonnées ───────────────────────────────────────────────
    meta_sheet = wb["SSSOM_Metadata"]
    metadata: dict[str, str] = {}
    for row in meta_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            metadata[str(row[0]).strip()] = str(row[1]).strip()

    curie_map: dict[str, str] = {}
    regular_meta: dict[str, str] = {}
    for k, v in metadata.items():
        if k.startswith("curie_map:"):
            curie_map[k.replace("curie_map:", "").strip()] = v
        else:
            regular_meta[k] = v

    # ── Lecture des mappings ──────────────────────────────────────────────────
    mapping_sheet = wb["Mappings"]
    rows = list(mapping_sheet.iter_rows(values_only=True))
    if not rows:
        raise ConversionError("La feuille 'Mappings' est vide.")

    header = [str(c).strip() for c in rows[0] if c is not None]
    mapping_rows = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        subject_id = str(row[0]).strip()
        if not subject_id.startswith("mainframe:"):
            continue
        mapping_rows.append(row)

    if not mapping_rows:
        raise ConversionError(
            "Aucun mapping valide trouvé. "
            "Vérifier que les subject_id commencent par 'mainframe:'."
        )

    # ── Écriture du fichier SSSOM/TSV ─────────────────────────────────────────
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        # En-tête YAML
        f.write(f"# mapping_set_id: {regular_meta.get('mapping_set_id', '')}\n")
        f.write(f"# mapping_set_description: \"{regular_meta.get('mapping_set_description', '')}\"\n")
        f.write(f"# license: {regular_meta.get('license', '')}\n")
        f.write(f"# mapping_date: {regular_meta.get('mapping_date', '')}\n")
        f.write(f"# subject_source: {regular_meta.get('subject_source', '')}\n")
        f.write(f"# subject_source_version: {regular_meta.get('subject_source_version', '')}\n")
        f.write(f"# object_source: {regular_meta.get('object_source', '')}\n")
        f.write(f"# object_source_version: {regular_meta.get('object_source_version', '')}\n")
        f.write(f"# mapping_tool: {regular_meta.get('mapping_tool', '')}\n")
        f.write("# curie_map:\n")
        for prefix, uri in curie_map.items():
            f.write(f"#   {prefix}: \"{uri}\"\n")
        f.write("#\n")

        # Données TSV
        f.write("\t".join(header) + "\n")
        for row in mapping_rows:
            values = []
            for i, col in enumerate(header):
                val = row[i] if i < len(row) and row[i] is not None else ""
                if col == "confidence" and val != "":
                    try:
                        val = f"{float(val):.1f}"
                    except (ValueError, TypeError):
                        val = str(val)
                else:
                    val = str(val).strip()
                values.append(val)
            f.write("\t".join(values) + "\n")

    logger.info(
        "Conversion SSSOM terminée : %d mappings écrits dans %s",
        len(mapping_rows), output_path
    )

    # ── Résumé ────────────────────────────────────────────────────────────────
    preview = []
    for row in mapping_rows[:5]:
        preview.append({
            "subject_id":  str(row[0]).strip() if row[0] else "",
            "predicate_id": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            "object_id":   str(row[4]).strip() if len(row) > 4 and row[4] else "",
        })

    return {
        "output_path":      str(output_path),
        "metadata_entries": len(regular_meta),
        "curie_prefixes":   len(curie_map),
        "mapping_rows":     len(mapping_rows),
        "mapping_set_id":   regular_meta.get("mapping_set_id", ""),
        "mapping_date":     regular_meta.get("mapping_date", ""),
        "preview":          preview,
    }
